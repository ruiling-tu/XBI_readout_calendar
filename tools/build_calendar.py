#!/usr/bin/env python3
"""Build an ICS calendar of upcoming clinical readouts for XBI holdings.

Outputs:
  - calendar/biotech-readouts.ics
  - data/events.json
"""
from __future__ import annotations

import csv
import datetime as dt
import json
import os
import re
import sys
import textwrap
import time
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
CACHE_DIR = DATA_DIR / "cache"
CTGOV_CACHE_DIR = CACHE_DIR / "ctgov"
PRICE_CACHE_DIR = CACHE_DIR / "prices"
CAL_DIR = BASE_DIR / "calendar"

XBI_HOLDINGS_URL = (
    "https://www.ssga.com/us/en/intermediary/etfs/library-content/products/"
    "fund-data/etfs/us/holdings-daily-us-en-xbi.xlsx"
)
CTGOV_SEARCH_URL = "https://clinicaltrials.gov/api/v2/studies"
CTGOV_STUDY_URL = "https://clinicaltrials.gov/api/v2/studies/{nct}"
CTGOV_VERSION_URL = "https://clinicaltrials.gov/api/v2/version"

STOOQ_QUOTE_URL = "https://stooq.com/q/l/?s={symbol}&f=sd2t2ohlcvn&h&e=csv"

WINDOW_MONTHS = 12
RECENT_RESULTS_DAYS = 30
CACHE_MAX_AGE_HOURS = 24

USER_AGENT = "biotech-readout-calendar/1.0 (+local script)"


@dataclass
class Company:
    name: str
    ticker: str


@dataclass
class TrialEvent:
    uid: str
    date: dt.date
    summary: str
    description: str
    url: str
    company: str
    ticker: str
    stock_price: Optional[float]
    stock_price_date: Optional[str]
    nct_id: str
    event_type: str  # readout_proxy or results_posted



def ensure_dirs() -> None:
    for d in [DATA_DIR, CACHE_DIR, CTGOV_CACHE_DIR, PRICE_CACHE_DIR, CAL_DIR]:
        d.mkdir(parents=True, exist_ok=True)


def now_utc() -> dt.datetime:
    return dt.datetime.now(dt.timezone.utc)


def http_get(url: str, cache_path: Optional[Path] = None, max_age_hours: int = 0) -> bytes:
    if cache_path and cache_path.exists() and max_age_hours > 0:
        age = now_utc() - dt.datetime.fromtimestamp(cache_path.stat().st_mtime, tz=dt.timezone.utc)
        if age.total_seconds() < max_age_hours * 3600:
            return cache_path.read_bytes()

    req = Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urlopen(req, timeout=30) as resp:
            data = resp.read()
    except HTTPError as e:
        raise RuntimeError(f"HTTP error {e.code} for {url}") from e
    except URLError as e:
        raise RuntimeError(f"Network error for {url}: {e}") from e

    if cache_path:
        cache_path.write_bytes(data)
    return data


def parse_partial_date(value: Optional[str]) -> Optional[dt.date]:
    if not value:
        return None
    value = value.strip()

    # YYYY-MM-DD
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", value)
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    # YYYY-MM
    m = re.match(r"^(\d{4})-(\d{2})$", value)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        last_day = (dt.date(year + (month // 12), (month % 12) + 1, 1) - dt.timedelta(days=1)).day
        return dt.date(year, month, last_day)

    # YYYY
    m = re.match(r"^(\d{4})$", value)
    if m:
        year = int(m.group(1))
        return dt.date(year, 12, 31)

    # YYYY Q[1-4] or YYYYQ[1-4]
    m = re.match(r"^(\d{4})\s*Q([1-4])$", value, re.IGNORECASE)
    if m:
        year = int(m.group(1))
        q = int(m.group(2))
        month = q * 3
        last_day = (dt.date(year + (month // 12), (month % 12) + 1, 1) - dt.timedelta(days=1)).day
        return dt.date(year, month, last_day)

    # Mon YYYY (e.g., Jan 2026)
    m = re.match(r"^([A-Za-z]{3})\s+(\d{4})$", value)
    if m:
        month_str = m.group(1).lower()
        year = int(m.group(2))
        months = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                  "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
        if month_str in months:
            month = months[month_str]
            last_day = (dt.date(year + (month // 12), (month % 12) + 1, 1) - dt.timedelta(days=1)).day
            return dt.date(year, month, last_day)

    return None


def load_overrides() -> Dict[str, Any]:
    overrides_path = DATA_DIR / "overrides.yaml"
    if not overrides_path.exists():
        return {}
    try:
        import yaml  # type: ignore
    except Exception:
        print("Note: overrides.yaml present but PyYAML is not installed; skipping overrides.")
        return {}
    with overrides_path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def parse_xbi_holdings_xlsx(xlsx_bytes: bytes) -> List[Company]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required to parse the XBI holdings xlsx. "
            "Install with: pip install -r requirements.txt"
        ) from e

    wb = load_workbook(filename=BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    header_row_idx = None
    header_map: Dict[str, int] = {}

    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if not row:
            continue
        lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if "ticker" in lower and ("name" in lower or "security name" in lower or "description" in lower):
            header_row_idx = idx
            for col_idx, val in enumerate(lower):
                header_map[val] = col_idx
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find header row in XBI holdings file.")

    def col_index(*candidates: str) -> Optional[int]:
        for cand in candidates:
            if cand in header_map:
                return header_map[cand]
        return None

    ticker_idx = col_index("ticker", "symbol")
    name_idx = col_index("name", "security name", "description")

    if ticker_idx is None or name_idx is None:
        raise RuntimeError("Could not find ticker/name columns in XBI holdings file.")

    companies: List[Company] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row:
            continue
        ticker = row[ticker_idx] if ticker_idx < len(row) else None
        name = row[name_idx] if name_idx < len(row) else None
        if not ticker or not name:
            continue
        ticker_str = str(ticker).strip().upper()
        name_str = str(name).strip()
        if not ticker_str or ticker_str == "N/A":
            continue
        companies.append(Company(name=name_str, ticker=ticker_str))

    return companies


def fetch_xbi_holdings() -> List[Company]:
    cache_path = CACHE_DIR / "xbi_holdings.xlsx"
    data = http_get(XBI_HOLDINGS_URL, cache_path=cache_path, max_age_hours=CACHE_MAX_AGE_HOURS)
    return parse_xbi_holdings_xlsx(data)


def fetch_stock_price(ticker: str) -> Tuple[Optional[float], Optional[str]]:
    cache_path = PRICE_CACHE_DIR / f"{ticker}.json"
    if cache_path.exists():
        age = now_utc() - dt.datetime.fromtimestamp(cache_path.stat().st_mtime, tz=dt.timezone.utc)
        if age.total_seconds() < CACHE_MAX_AGE_HOURS * 3600:
            try:
                cached = json.loads(cache_path.read_text(encoding="utf-8"))
                return cached.get("price"), cached.get("date")
            except Exception:
                pass

    symbol = f"{ticker.lower()}.us"
    url = STOOQ_QUOTE_URL.format(symbol=symbol)
    raw = http_get(url)
    text = raw.decode("utf-8", errors="replace").strip()
    reader = csv.DictReader(text.splitlines())
    for row in reader:
        close_str = row.get("Close") or row.get("close")
        date_str = row.get("Date") or row.get("date")
        if close_str and close_str not in {"N/A", ""}:
            try:
                price = float(close_str)
            except ValueError:
                price = None
            result = {"price": price, "date": date_str}
            cache_path.write_text(json.dumps(result), encoding="utf-8")
            return price, date_str

    cache_path.write_text(json.dumps({"price": None, "date": None}), encoding="utf-8")
    return None, None


def ctgov_search(company: Company, start_date: dt.date, end_date: dt.date) -> List[str]:
    params = {
        "query.spons": company.name,
        "filter.advanced": f"AREA[PrimaryCompletionDate]RANGE[{start_date},{end_date}]",
        "pageSize": 100,
        "countTotal": "true",
    }

    nct_ids: List[str] = []
    next_page_token: Optional[str] = None

    while True:
        if next_page_token:
            params["pageToken"] = next_page_token
        url = CTGOV_SEARCH_URL + "?" + urlencode(params)
        cache_path = CTGOV_CACHE_DIR / f"search_{company.ticker}.json"
        data = http_get(url, cache_path=cache_path, max_age_hours=CACHE_MAX_AGE_HOURS)
        payload = json.loads(data.decode("utf-8"))
        for study in payload.get("studies", []):
            nct = study.get("protocolSection", {}).get("identificationModule", {}).get("nctId")
            if nct:
                nct_ids.append(nct)
        next_page_token = payload.get("nextPageToken")
        if not next_page_token:
            break

    return sorted(set(nct_ids))


def ctgov_fetch_study(nct_id: str) -> Dict[str, Any]:
    cache_path = CTGOV_CACHE_DIR / f"{nct_id}.json"
    data = http_get(CTGOV_STUDY_URL.format(nct=nct_id), cache_path=cache_path, max_age_hours=CACHE_MAX_AGE_HOURS)
    return json.loads(data.decode("utf-8"))


def strip_markdown(text: str) -> str:
    text = re.sub(r"`+", "", text)
    text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
    text = re.sub(r"\*([^*]+)\*", r"\1", text)
    text = re.sub(r"_([^_]+)_", r"\1", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def infer_moa(drug_names: List[str], summary: str) -> Optional[str]:
    text = (" ".join(drug_names) + " " + summary).lower()
    if "car-t" in text or "car t" in text:
        return "CAR-T cell therapy"
    if "gene therapy" in text:
        return "Gene therapy"
    if "mrna" in text or "mRNA" in text:
        return "mRNA-based therapeutic"
    if "monoclonal" in text or any(name.lower().endswith("mab") for name in drug_names):
        return "Monoclonal antibody"
    if "siRNA" in text or "sirna" in text:
        return "siRNA therapeutic"
    if "small molecule" in text:
        return "Small-molecule therapeutic"
    return None


def safe_get(obj: Dict[str, Any], path: List[str]) -> Any:
    cur: Any = obj
    for key in path:
        if not isinstance(cur, dict) or key not in cur:
            return None
        cur = cur[key]
    return cur


def build_event(company: Company, study: Dict[str, Any], stock_price: Optional[float], stock_date: Optional[str],
                overrides: Dict[str, Any], event_type: str, event_date: dt.date) -> TrialEvent:
    protocol = study.get("protocolSection", {})
    status = protocol.get("statusModule", {})
    design = protocol.get("designModule", {})
    outcomes = protocol.get("outcomesModule", {})
    eligibility = protocol.get("eligibilityModule", {})

    nct_id = safe_get(protocol, ["identificationModule", "nctId"]) or study.get("nctId")
    brief_title = safe_get(protocol, ["identificationModule", "briefTitle"]) or "Clinical study"
    url = f"https://clinicaltrials.gov/study/{nct_id}"

    drug_names: List[str] = []
    for intr in design.get("interventions", []) or []:
        intr_type = (intr.get("type") or "").lower()
        name = intr.get("name") or ""
        if intr_type in {"drug", "biological", "genetic", "cell"} and name:
            drug_names.append(name)
    if not drug_names:
        for intr in design.get("interventions", []) or []:
            name = intr.get("name") or ""
            if name:
                drug_names.append(name)

    phase = ", ".join(design.get("phases", []) or []) or "Not specified"
    conditions = protocol.get("conditionsModule", {}).get("conditions", []) or []

    summary_text = protocol.get("descriptionModule", {}).get("briefSummary") or ""
    summary_text = strip_markdown(summary_text) if summary_text else ""

    primary_outcomes = outcomes.get("primaryOutcomes", []) or []
    secondary_outcomes = outcomes.get("secondaryOutcomes", []) or []

    def outcome_lines(items: List[Dict[str, Any]]) -> List[str]:
        lines = []
        for item in items:
            measure = item.get("measure") or ""
            time_frame = item.get("timeFrame") or ""
            desc = item.get("description") or ""
            bits = [b for b in [measure, time_frame, desc] if b]
            if bits:
                lines.append(" — ".join(bits))
        return lines

    primary_lines = outcome_lines(primary_outcomes)
    secondary_lines = outcome_lines(secondary_outcomes)

    gender = eligibility.get("sex") or "Not specified"
    min_age = eligibility.get("minimumAge") or "Not specified"
    max_age = eligibility.get("maximumAge") or "Not specified"
    healthy_volunteers = eligibility.get("healthyVolunteers")
    pop_bits = [f"Sex: {gender}", f"Age: {min_age} to {max_age}"]
    if healthy_volunteers is not None:
        pop_bits.append(f"Healthy Volunteers: {healthy_volunteers}")
    population = "; ".join(pop_bits)

    criteria = eligibility.get("eligibilityCriteria")
    criteria = strip_markdown(criteria) if criteria else ""
    if criteria:
        criteria = textwrap.shorten(criteria, width=500, placeholder="...")

    drug = ", ".join(drug_names) if drug_names else "Not specified"
    moa = infer_moa(drug_names, summary_text) or "Not specified in ClinicalTrials.gov"

    override = overrides.get(nct_id, {}) if nct_id else {}
    if override:
        drug = override.get("drug", drug)
        moa = override.get("moa", moa)
        if override.get("indication"):
            conditions = [override["indication"]]
        if override.get("patient_population"):
            population = override["patient_population"]
        if override.get("endpoints"):
            primary_lines = override["endpoints"].get("primary", primary_lines)
            secondary_lines = override["endpoints"].get("secondary", secondary_lines)

    indication = "; ".join(conditions) if conditions else "Not specified"

    stock_line = "Not available"
    if stock_price is not None:
        stock_line = f"${stock_price:.2f} (as of {stock_date})" if stock_date else f"${stock_price:.2f}"

    title_drug = drug if drug != "Not specified" else brief_title
    summary = f"{company.name} ({company.ticker}) — {title_drug} — {event_type.replace('_', ' ').title()}"

    desc_lines = [
        f"Company: {company.name}",
        f"Ticker: {company.ticker}",
        f"Stock Price: {stock_line}",
        f"Drug: {drug}",
        f"MOA: {moa}",
        f"Phase: {phase}",
        f"Indication: {indication}",
        f"Patient Population: {population}",
    ]

    if primary_lines:
        desc_lines.append("Primary Endpoints:")
        desc_lines.extend([f"- {line}" for line in primary_lines])
    if secondary_lines:
        desc_lines.append("Secondary Endpoints:")
        desc_lines.extend([f"- {line}" for line in secondary_lines])

    if criteria:
        desc_lines.append(f"Eligibility Criteria (excerpt): {criteria}")

    if override.get("endpoint_notes"):
        desc_lines.append(f"Endpoint Notes: {override['endpoint_notes']}")

    if summary_text:
        desc_lines.append(f"Summary: {summary_text}")

    desc_lines.append(f"ClinicalTrials.gov: {url}")

    description = "\n".join(desc_lines)

    uid = f"{nct_id}-{event_type}-{event_date.isoformat()}"

    return TrialEvent(
        uid=uid,
        date=event_date,
        summary=summary,
        description=description,
        url=url,
        company=company.name,
        ticker=company.ticker,
        stock_price=stock_price,
        stock_price_date=stock_date,
        nct_id=nct_id,
        event_type=event_type,
    )


def ics_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def fold_ics_line(line: str) -> str:
    if len(line) <= 75:
        return line
    parts = [line[:75]]
    line = line[75:]
    while line:
        parts.append(" " + line[:74])
        line = line[74:]
    return "\r\n".join(parts)


def build_ics(events: List[TrialEvent]) -> str:
    now = now_utc().strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//biotech-readout-calendar//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
    ]
    for ev in events:
        dtstart = ev.date.strftime("%Y%m%d")
        dtend = (ev.date + dt.timedelta(days=1)).strftime("%Y%m%d")
        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{ics_escape(ev.uid)}",
            f"DTSTAMP:{now}",
            f"DTSTART;VALUE=DATE:{dtstart}",
            f"DTEND;VALUE=DATE:{dtend}",
            f"SUMMARY:{ics_escape(ev.summary)}",
            f"DESCRIPTION:{ics_escape(ev.description)}",
            f"URL:{ics_escape(ev.url)}",
            "END:VEVENT",
        ])
    lines.append("END:VCALENDAR")

    folded = [fold_ics_line(line) for line in lines]
    return "\r\n".join(folded) + "\r\n"


def load_ctgov_version() -> Optional[str]:
    try:
        raw = http_get(CTGOV_VERSION_URL, cache_path=CACHE_DIR / "ctgov_version.json", max_age_hours=12)
        data = json.loads(raw.decode("utf-8"))
        return data.get("dataTimestamp")
    except Exception:
        return None


def main() -> int:
    ensure_dirs()
    overrides = load_overrides()

    today = dt.date.today()
    end_date = today + dt.timedelta(days=WINDOW_MONTHS * 30)

    print("Fetching XBI holdings...")
    companies = fetch_xbi_holdings()

    print(f"XBI holdings: {len(companies)} companies")

    ctgov_timestamp = load_ctgov_version()

    events: List[TrialEvent] = []
    for idx, company in enumerate(companies, start=1):
        print(f"[{idx}/{len(companies)}] {company.ticker} {company.name}")
        price, price_date = fetch_stock_price(company.ticker)

        try:
            nct_ids = ctgov_search(company, today, end_date)
        except Exception as e:
            print(f"  Warning: CT.gov search failed for {company.ticker}: {e}")
            continue

        for nct_id in nct_ids:
            try:
                study = ctgov_fetch_study(nct_id)
            except Exception as e:
                print(f"  Warning: CT.gov fetch failed for {nct_id}: {e}")
                continue

            protocol = study.get("protocolSection", {})
            status = protocol.get("statusModule", {})
            primary_date = safe_get(status, ["primaryCompletionDateStruct", "date"])
            primary_date = primary_date or status.get("primaryCompletionDate")
            primary_date = parse_partial_date(primary_date)

            if primary_date and today <= primary_date <= end_date:
                events.append(build_event(
                    company, study, price, price_date, overrides,
                    event_type="readout_proxy",
                    event_date=primary_date,
                ))

            results_date = safe_get(status, ["resultsFirstPostDateStruct", "date"]) or status.get("resultsFirstPostDate")
            results_date = parse_partial_date(results_date)
            if results_date:
                days_ago = (today - results_date).days
                if 0 <= days_ago <= RECENT_RESULTS_DAYS:
                    events.append(build_event(
                        company, study, price, price_date, overrides,
                        event_type="results_posted",
                        event_date=results_date,
                    ))

    events.sort(key=lambda e: (e.date, e.company, e.nct_id))

    ics_text = build_ics(events)
    ics_path = CAL_DIR / "biotech-readouts.ics"
    ics_path.write_text(ics_text, encoding="utf-8")

    events_json = [
        {
            "uid": e.uid,
            "date": e.date.isoformat(),
            "summary": e.summary,
            "description": e.description,
            "url": e.url,
            "company": e.company,
            "ticker": e.ticker,
            "stock_price": e.stock_price,
            "stock_price_date": e.stock_price_date,
            "nct_id": e.nct_id,
            "event_type": e.event_type,
        }
        for e in events
    ]

    output = {
        "generated_at": now_utc().isoformat(),
        "ctgov_data_timestamp": ctgov_timestamp,
        "window_start": today.isoformat(),
        "window_end": end_date.isoformat(),
        "events": events_json,
    }
    (DATA_DIR / "events.json").write_text(json.dumps(output, indent=2), encoding="utf-8")

    print(f"Wrote {len(events)} events to {ics_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
